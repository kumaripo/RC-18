import openpyxl
import jira.client
from jira.client import JIRA

SheetName = 'Jira_Details'
wb=openpyxl.Workbook()
ws = wb.active
ws.title = SheetName
ws.cell(row=1,column = 1).value = 'Sl No.'
ws.cell(row=1,column = 2).value = 'PBI Number'
ws.cell(row=1,column = 3).value = 'Issue Type'
ws.cell(row=1,column = 4).value = 'Sprint'
ws.cell(row=1,column = 5).value = 'Status'
ws.cell(row=1,column = 6).value = 'Labels'
ws.cell(row=1,column = 7).value = 'Feature'
ws.cell(row=1,column = 8).value = 'Priority'
ws.cell(row=1,column = 9).value = 'Summary'
ws.cell(row=1,column = 10).value = 'TCs Identified'
ws.cell(row=1,column = 11).value = 'TCs Coded'
ws.cell(row=1,column = 12).value = 'Reporter'
ws.cell(row=1,column = 13).value = 'Assignee'
ws.cell(row=1,column = 14).value = 'Estimated Time'
ws.cell(row=1,column = 15).value = 'Remaining TIme'
ws.cell(row=1,column = 16).value = 'Time Spent'
rowIndex = 2


options = {'server': 'http://172.27.39.6:8071', 'verify':False}
jira = JIRA(options, basic_auth=('dineshvijay.9140', 'indrani@9140'))
user = ['project=RadisysCodeathon2k18 AND reporter=dineshvijay.9140','project=RadisysCodeathon2k18 AND reporter=abdulla']
for p in user:
    issues_in_project = jira.search_issues(p)
    for value in issues_in_project:
        PBI_key = str(value)
        issue_type = value.fields.issuetype
        try:
            sprint = value.fields.customfield_10004[0].split(',')[3].split("=")[1]
        except:
            sprint = "NA"
        status = value.fields.status
        labels = "NA"
        feature = "NA"
        Labels = (getattr(value.fields(), 'labels'))
        if Labels != None:
            if len(value.fields.labels) > 1:
                j = 0 
                for i in value.fields.labels:
                    if len(i) > 3:
                        feature = value.fields.labels[j]
                    else:
                        labels = value.fields.labels[j]
                    j = j+1
            elif len(value.fields.labels) == 1:
                 feature = "NA"
                 labels = value.fields.labels[0]
        else:
            feature = "NA"
            labels  = "NA"
        priority    = value.fields.priority
        summary     = value.fields.summary
        tc_ident = (getattr(value.fields(), 'customfield_10008'))
        if tc_ident == None:
            tc_ident    = "NA"
        else:
            tc_ident    = value.fields.customfield_10008
        tc_coded = (getattr(value.fields(), 'customfield_10009'))
        if tc_coded == None:
            tc_coded    = "NA"
        else:
            tc_coded    = value.fields.customfield_10009
        reporter    = value.fields.reporter
        assignee    = value.fields.assignee
        est_time = (getattr(value.fields(), 'timeoriginalestimate'))
        if est_time == None:
            est_time = 0
        else:
            est_time = est_time/3600
        rem_time = (getattr(value.fields(), 'timeestimate'))
        if rem_time == None:
            rem_time = 0
        else:
            rem_time = rem_time/3600
        time_spent = (getattr(value.fields(), 'timespent'))
        if time_spent == None:
            time_spent = 0
        else:
            time_spent = time_spent/3600
       
        ws.cell(row=rowIndex, column=1).value = rowIndex-1
        ws.cell(row=rowIndex, column=2).value = str(PBI_key)
        ws.cell(row=rowIndex, column=3).value = str(issue_type)
        ws.cell(row=rowIndex, column=4).value = str(sprint)
        ws.cell(row=rowIndex, column=5).value = str(status)
        ws.cell(row=rowIndex, column=6).value = str(labels)
        ws.cell(row=rowIndex, column=7).value = str(feature)
        ws.cell(row=rowIndex, column=8).value = str(priority)
        ws.cell(row=rowIndex, column=9).value = str(summary)
        ws.cell(row=rowIndex, column=10).value = str(tc_ident)
        ws.cell(row=rowIndex, column=11).value = str(tc_coded)
        ws.cell(row=rowIndex, column=12).value = str(reporter)
        ws.cell(row=rowIndex, column=13).value = str(assignee)
        ws.cell(row=rowIndex, column=14).value = est_time
        ws.cell(row=rowIndex, column=15).value = rem_time
        ws.cell(row=rowIndex, column=16).value = time_spent
        rowIndex = rowIndex + 1
wb.save('Jira_Details.xlsx')



